"""
MADE BY PRATHAM GUPTA (https://github.com/PrathamGupta06/cbse-results-analyzer)
"""

import sys
import pandas as pd
from openpyxl import Workbook, load_workbook
from _functions import *

# TODO:
#     Fix for case where student is absent

print("-"*64)
print("Welcome to CBSE Results Analyzer.\nThis Program is made by Pratham Gupta ("
      "https://github.com/PrathamGupta06/cbse-results-analyzer)")
print("-"*64)

if len(sys.argv) > 1:
    # Command Line Argument Mode
    input_file = sys.argv[1]
    if not validate_input_path(input_file):
        print("ERROR: Invalid Input File Path")
        exit()

    output_path_excel = sys.argv[2]
    if not validate_output_path(output_path_excel):
        print("ERROR: Invalid Output Path")
        exit()
    if not output_path_excel.endswith(".xlsx"):   # Append .xlsx if not present
        output_path_excel += ".xlsx"

    mode = sys.argv[3]
    if mode not in ("10th", "12th"):
        print("ERROR: Invalid Class")
        print("Valid Classes: 10th, 12th")
else:
    # Taking input from user
    input_file = input("\nEnter Input Result File Path (e.g demo/12th.txt): ")
    if not validate_input_path(input_file):
        print("ERROR: Invalid Input File Path")
        exit()

    output_path_excel = input("Enter Output Path (e.g 12th.xlsx): ")
    if not validate_output_path(output_path_excel):
        print("ERROR: Invalid Output Path")
        exit()
    if not output_path_excel.endswith(".xlsx"):   # Append .xlsx if not present
        output_path_excel += ".xlsx"

    mode = input("Enter Class (10th or 12th): ")
    if mode not in ("10th", "12th"):
        print("ERROR: Invalid Class")
        print("Valid Classes: 10th, 12th")

# Create Workbook
wb = Workbook()
save_wb(wb, output_path_excel)
print("Blank Workbook Created")
wb.close()

column_widths = {
    "Roll No": 20,
    "Gender": 8,
    "Name": 20,
    "Subject": 15,
    "Grade": 6,
    "GR1": 5,
    "GR2": 5,
    "GR3": 5,
    "Result": 8,
    "Best 5 Percentage": 10,
    "Compartment Subject": 15
}

# Get the data from the file
headers, data, AllSubjectNames = get_data(input_file, mode)
main_df = pd.DataFrame(data, columns=headers)

# Convert Pandas dataframe to Excel
wb = load_workbook(output_path_excel)
main_sheet = wb["Sheet"]
main_sheet.title = "Main Result Sheet"
write_data(main_sheet, main_df)
adjust_column_widths(main_sheet, column_widths)

# Freezing the First Row and First Three Columns of Main Sheet
main_sheet.freeze_panes = "D2"

# Creating Individual Subject Sheets
for SubjectName in AllSubjectNames:
    index_no = main_df.columns.get_loc(SubjectName)
    subject_df = main_df.iloc[:, [0, 1, 2, index_no, index_no + 1]]  # filter the columns to the subject name and marks
    subject_df = subject_df[subject_df[SubjectName] >= 0]  # filter the rows which do not have subject marks
    subject_sheet = wb.create_sheet(SubjectName)
    write_data(subject_sheet, subject_df)
    adjust_column_widths(subject_sheet, column_widths)

# Analyzing the Results

# Top 5 Analyze
# Splitting the data into Male and Female
male = [student for student in data if student[1] == "M"]
female = [student for student in data if student[1] == "F"]

# Sorting the Male and Female list by Best 5 Percentage
top_n_children = 5
best_5_index = headers.index("Best 5 percentage")
top_male = sorted(male, key=lambda student: student[best_5_index], reverse=True)[:top_n_children]
top_female = sorted(female, key=lambda student: student[best_5_index], reverse=True)[:top_n_children]

# Removing the Unnecessary Columns in the list
top_male = [[student[0], student[1], student[2], student[best_5_index]] for student in top_male]
top_female = [[student[0], student[1], student[2], student[best_5_index]] for student in top_female]

# Analyzation related to marks
children_with_full_marks = [["Roll No", "Gender", "Name", "Subject"]]
total_distinctions = 0
all_5_distinctions = 0  # All 5 Subjects Distinction Analyze
all_5_distinction_students = []  # All 5 Subjects Distinction Student Details

for student_data in data:
    student_distinctions = 0  # Individual Student Distinctions
    for index, value in enumerate(student_data):
        if type(value) is int or type(value) is float:
            if value >= 75:
                student_distinctions += 1
                total_distinctions += 1
            if value == 100:
                children_with_full_marks.append([student_data[0], student_data[1], student_data[2], headers[index]])
    #  For All 5 Distinctions
    if student_distinctions >= 5:
        all_5_distinctions += 1
        all_5_distinction_students.append([student_data[0], student_data[1], student_data[2]])

# Writing the analysis data to worksheet
wb.create_sheet('Analysis', 1)
analyze_ws = wb['Analysis']

# Children with Full Marks
title = "Children With Full Marks"
append_title(analyze_ws, title, end_column=4)
for i in children_with_full_marks:
    analyze_ws.append(i)


# Overall Toppers
title = "Overall {} Toppers".format(top_n_children)
append_title(analyze_ws, title, end_column=4)

# Male Toppers
append_title(analyze_ws, title="Male", end_column=4)
analyze_ws.append(["Roll No", "Gender", "Name", "Best 5 Percentage"])
for i in top_male:
    analyze_ws.append(i)

# Female Toppers
append_title(analyze_ws, title="Female", end_column=4)
analyze_ws.append(["Roll No", "Gender", "Name", "Best 5 Percentage"])
for i in top_female:
    analyze_ws.append(i)

# Distinctions in all subjects
title = "Total Distinctions: {0}".format(total_distinctions)
append_title(analyze_ws, title, end_column=4)

# Distinctions in all 5 Subjects
title = "Total Distinctions in all 5 subjects: {}".format(all_5_distinctions)
append_title(analyze_ws, title, end_column=4)

# analyze_ws.append(["Distinctions in all 5 Subjects Students"])
# analyze_ws.cell(row=ws.max_row), column=1).font = Font(bold=True)
# analyze_ws.merge_cells(start_row=ws.max_row), start_column=1, end_row=ws.max_row), end_column=3)
# for i in all_5_distinction_students:
#     analyze_ws.append(i)
# last_row += len(all_5_distinction_students)

# Adjust the widths
analyze_ws.column_dimensions["A"].width = column_widths["Roll No"]
analyze_ws.column_dimensions["B"].width = column_widths["Gender"]
analyze_ws.column_dimensions["C"].width = column_widths["Name"]
analyze_ws.column_dimensions["D"].width = column_widths["Subject"]

# Save Workbook
save_wb(wb, output_path_excel)
wb.close()
print("Workbook Saved at", os.path.abspath(output_path_excel))

# Launch Workbook if mode is not Command Line Argument
if len(sys.argv) == 1:
    os.startfile(os.path.abspath(output_path_excel))
print("Program Completed Successfully")
