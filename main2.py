import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter
CBSE_CLASS_12_SUBJECT_CODES = {
    # Language Subjects
    "001": "ENGLISH ELECTIVE",
    "301": "ENGLISH CORE",
    "002": "HINDI ELECTIVE",
    "302": "HINDI CORE",
    "003": "URDU ELECTIVE",
    "303": "URDU CORE",
    "022": "SANSKRIT ELECTIVE",
    "322": "SANSKRIT CORE",
    "104": "PUNJABI",
    "105": "BENGALI",
    "106": "TAMIL",
    "107": "TELUGU",
    "108": "SINDHI",
    "109": "MARATHI",
    "110": "GUJARATI",
    "111": "MANIPURI",
    "112": "MALAYALAM",
    "113": "ODIA",
    "114": "ASSAMESE",
    "115": "KANNADA",
    "116": "ARABIC",
    "117": "TIBETAN",
    "118": "FRENCH",
    "120": "GERMAN",
    "121": "RUSSIAN",
    "123": "PERSIAN",
    "124": "NEPALI",
    "125": "LIMBOO",
    "126": "LEPCHA",
    "189": "TELUGU TELANGANA",
    "192": "BODO",
    "193": "TANGKHUL",
    "194": "JAPANESE",
    "195": "BHUTIA",
    "196": "SPANISH",
    "197": "KASHMIRI",
    "198": "MIZO",
    "199": "BAHASA MELAYU",

    # Academic Subjects
    "027": "HISTORY",
    "028": "POLITICAL SCIENCE",
    "029": "GEOGRAPHY",
    "030": "ECONOMICS",
    "031": "CARNATIC MUSIC (VOCAL)",
    "032": "CARNATIC MUSIC( MELODIC INSTRUMENTS).",
    "033": "CARNATIC MUSIC ( PERCUSSION INSTRUMENTS - MRIDANGAM)",
    "034": "HINDUSTANI MUSIC (VOCAL)",
    "035": "HINDUSTANI MUSIC ( MELODIC INSTRUMENTS).",
    "036": "HINDUSTANI MUSIC ( PERCUSSION INSTRUMENTS)",
    "037": "PSYCHOLOGY",
    "039": "SOCIOLOGY",
    "041": "MATHEMATICS",
    "042": "PHYSICS",
    "043": "CHEMISTRY",
    "044": "BIOLOGY",
    "045": "BIOTECHNOLOGY",
    "046": "ENGINEERING GRAPHICS",
    "048": "PHYSICAL EDUCATION",
    "049": "PAINTING",
    "050": "GRAPHICS",
    "051": "SCULPTURE",
    "052": "APPLIED/ COMMERCIAL ART",
    "054": "BUSINESS STUDIES",
    "055": "ACCOUNTANCY",
    "056": "KATHAK - DANCE",
    "057": "BHARATNATYAM - DANCE",
    "058": "KUCHIPUDI-DANCE",
    "059": "ODISSI - DANCE",
    "060": "MANIPURI - DANCE",
    "061": "KATHAKALI - DANCE",
    "064": "HOME SCIENCE",
    "265": "INFORMATICS PRACTICES (OLD)(Only for XII)",
    "065": "INFORMATICS PRACTICES (NEW)",
    "283": "COMPUTER SCIENCE (OLD) (Only for XII)",
    "083": "COMPUTER SCIENCE (NEW)",
    "066": "ENTREPRENEURSHIP",
    "073": "KNOWLEDGE TRADITION & PRACTICES OF INDIA",
    "074": "LEGAL STUDIES",
    "076": "NATIONAL CADET CORPS (NCC)",
}

CBSE_CLASS_10_SUBJECT_CODES = {
    # Language Subjects
    "002": "HINDI COURSE-A",
    "085": "HINDI COURSE-B",
    "184": "ENGLISH LANG & LIT.",
    "003": "URDU COURSE-A",
    "303": "URDU COURSE-B",
    "004": "PUNJABI",
    "005": "BENGALI",
    "006": "TAMIL",
    "007": "TELUGU",
    "008": "SINDHI",
    "009": "MARATHI",
    "010": "GUJARATI",
    "011": "MANIPURI",
    "012": "MALAYALAM",
    "013": "ODIA",
    "014": "ASSAMESE",
    "015": "KANNADA",
    "016": "ARABIC",
    "017": "TIBETAN",
    "018": "FRENCH",
    "020": "GERMAN",
    "021": "RUSSIAN",
    "023": "PERSIAN",
    "024": "NEPALI",
    "025": "LIMBOO",
    "026": "LEPCHA",
    "089": "TELUGU TELANGANA",
    "092": "BODO",
    "093": "TANGKHUL",
    "094": "JAPANESE",
    "095": "BHUTIA",
    "096": "SPANISH",
    "097": "KASHMIRI",
    "098": "MIZO",
    "099": "BAH ASA MELAYU",
    "122": "SANSKRIT",
    "131": "RAI",
    "132": "GURUNG",
    "133": "TAMANG",
    "134": "SHERPA",
    "136": "THAI",

    # Academic Subjects
    "041": "MATHEMATICS -STANDARD",
    "241": "MATHEMATICS -BASIC",
    "086": "SCIENCE",
    "087": "SOCIAL SCIENCE",

    # Additional Academic Subjects
    "031": "CARNATIC MUSIC (VOCAL)",
    "032": "CARNATIC MUSIC (MELODIC INSTRUMENTS)",
    "033": "CARNATIC MUSIC (PERCUSSION INSTRUMENTS)",
    "034": "HINDUSTANI MUSIC (VOCAL)",
    "035": "HINDUSTANI MUSIC (MELODIC INSTRUMENTS)",
    "036": "HINDUSTANI MUSIC (PERCUSSION INSTRUMENTS)",
    "049": "PAINTING",
    "064": "HOME SCIENCE",
    "076": "NATIONAL CADET CORPS (NCC)",
    "165": "COMPUTER APPLICATIONS",
    "154": "ELEMENTS OF BUSINESS",
    "254": "ELEMENTS OF BOOK KEEPING & ACCOUNTANCY",

    # Skill Subjects
    "401": "RETAILING",
    "402": "INFORMATION TECHNOLOGY",
    "403": "SECURITY",
    "404": "AUTOMOTIVE",
    "405": "INTRODUCTION TO FINANCIAL MARKETS",
    "406": "INTRODUCTION TO TOURISM",
    "407": "BEAUTY & WELLNESS",
    "408": "AGRICULTURE",
    "409": "FOOD PRODUCTION",
    "410": "FRONT OFFICE OPERATIONS",
    "411": "BANKING & INSURANCE",
    "412": "MARKETING & SALES",
    "413": "HEALTH CARE",
    "414": "APPAREL",
    "415": "MEDIA",
    "416": "MULTI SKILL FOUNDATION COURSE",
    "417": "ARTIFICIAL INTELLIGENCE",

}

ROLL_NUMBER_LENGTH = 8
NAME_STARTING = 13
NAME_ENDING = 64
SUBJECT_STARTING = 64
SUBJECT_ENDING = 113
GRADE_STARTING = 64
GRADE_ENDING = 113
RESULT_STARTING = 114
RESULT_ENDING = 121
# TODO: Complete the Groups
GROUPS = {
    "Science": ["PHYSICS", "CHEMISTRY", "MATHS"]

}


def is_integer(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def convert_subject_code_to_name(subject_codes_list):
    sub_names = []
    for subjectCode in subject_codes_list:
        if subjectCode in CBSE_CLASS_10_SUBJECT_CODES:
            sub_names.append(CBSE_CLASS_10_SUBJECT_CODES[subjectCode])
        else:
            sub_names.append(subjectCode)
    return sub_names


def get_subject_headings_and_remove_other_stuff(file_name, new_file_name="only_records.txt"):
    total_students = 0
    unique_subject_codes = []
    file = open(file_name, "r")
    new_file = open("output\\" + new_file_name, "w")

    # total_students - number of students in the file
    # unique_subject_codes - list of unique subject codes
    # file - original result file
    # file_new - new file without headings

    lines = file.readlines()
    i = 0
    while i <= len(lines) - 2:
        name_line = lines[i]
        if is_integer(name_line[:ROLL_NUMBER_LENGTH]):
            # Line contains student name

            total_students += 1
            sub_codes = name_line[SUBJECT_STARTING:SUBJECT_ENDING].split()
            # sub_codes - Subject Codes of Subjects given by student

            # writing student records to new file
            new_file.writelines([lines[i], lines[i + 1]])

            # Adding new subject codes to list of unique subject codes
            for sub_code in sub_codes:
                if sub_code not in unique_subject_codes:
                    unique_subject_codes.append(sub_code)
        i += 1

    unique_subject_names = convert_subject_code_to_name(unique_subject_codes)

    new_file.close()
    file.close()

    return unique_subject_names, unique_subject_codes


def arrange_records_in_order(header, sub_codes, record_line_1, record_line_2):
    roll_no = record_line_1[:ROLL_NUMBER_LENGTH]
    name = record_line_1[NAME_STARTING:NAME_ENDING]
    subject_codes_of_student = record_line_1[SUBJECT_STARTING:SUBJECT_ENDING].split()
    subject_marks_of_student = record_line_2[GRADE_STARTING:GRADE_ENDING].split()[::2]

    ordered_marks = [""] * len(sub_codes)

    # Arranging marks in order of subject codes given in sub_codes
    # In case of subject code not given by the student, it will be marked as ""
    # e.g
    # sub_codes = ['184', '122', '041', '086', '087', '085', '241', '165', '064']
    # subject_codes_of_student = ['184', '122', '041', '086', '087']
    # subject_marks_of_student = ['086', '090', '094', '091', '096']
    # then marks will be ordered as :-
    # ['086', '090', '094', '091', '096', '', '', '', '']
    # Here since the student doesn't have subject code '085', '241', '165', '064'
    # those subject marks are set to ""

    # This is done to make sure that while appending the marks to the Excel sheet,
    # the marks are in the same order as the subject codes given in the headers.

    for code, marks in zip(subject_codes_of_student, subject_marks_of_student):
        if is_integer(marks):
            marks = int(marks)
        ordered_marks[sub_codes.index(code)] = marks

    return [roll_no, name] + ordered_marks


def intitalize_workbook(workbook_path):
    # create a workbook at workbook path
    # create an all sheet and delete the sheet named "Sheet"
    # return the workbook object
    wb = Workbook()
    wb.create_sheet("all", 0)
    del wb["Sheet"]
    wb.save(workbook_path)


original_file_name = r"input\result_10th.txt"
subject_names, subject_codes = get_subject_headings_and_remove_other_stuff(original_file_name, "temp.txt")
headers = ["Roll Number", "Name"] + subject_names

f = open("output\\temp.txt", "r")

intitalize_workbook("output\\result_10th.xlsx")
workbook = load_workbook("output\\result_10th.xlsx")
ws = workbook["all"]
lines = f.readlines()

# add the records to a list data in correct order
data = []
ws.append(headers)
i = 0
while i <= len(lines) - 2:
    correct_record = arrange_records_in_order(headers, subject_codes, lines[i], lines[i + 1])
    ws.append(correct_record)
    i += 2

tab = Table(displayName="Table1", ref="A1:" + get_column_letter(len(headers)) + str(ws.max_row))

style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
ws.add_table(tab)

workbook.save("output\\result_10th.xlsx")
f.close()
os.remove("output\\temp.txt")


# subjects_regex = (?<=\s)\d\d\d(?=\s)
# roll_no_regex = \d{8,}
# gender_regex = [FM]
# name_regex = [A-Z]{2,} then remove "PASS"
# marks_regex = (?<=\s)\d\d\d(?=\s)
# pass_fail_regex = PASS|FAIL
# grades_regex = [ABCDE][12]|(?<=\s)[E](?=\s)
# create a data list with [roll no, gender, name, marks, grade, pass/fail, best 5 subject marks]